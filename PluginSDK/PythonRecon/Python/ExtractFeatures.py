from radiomics import featureextractor
import h5py
import SimpleITK as sitk
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import numbers

import numpy as np
import nipy
import os
import ReadNiiPathFiles
import re
import xlrd

open_label  = xlrd.open_workbook('D:\数据\MICCAI_BraTS17_Data_Training_for_TGuo_2017modified\Brats17TrainingData\survival_data.xlsx')
labels = open_label.sheet_by_name('marked')
file_name_list = labels.col_values(0)[1:]

## read nii file
# HGG labeled data
hgg_folder_path = 'D:\数据\MICCAI_BraTS17_Data_Training_for_TGuo_2017modified\Brats17TrainingData\HGG'
hgg_file_path_list = ReadNiiPathFiles.walk_dir(hgg_folder_path, topdown=True)
flair_re = re.compile(r'flair.nii')
seg_re = re.compile(r'seg.nii')
t1_re = re.compile(r't1.nii')
t1ce_re = re.compile(r't1ce.nii')
t2_re = re.compile(r't2.nii')
hgg_flair_list = []
hgg_seg_list = []
hgg_t1_list = []
hgg_t1ce_list =[]
hgg_t2_list = []
for x in hgg_file_path_list:
    if re.search(flair_re,x):
        flair_file = nipy.load_image(x)
        hgg_flair_list.append(np.array(flair_file.get_data()).tolist())
    if re.search(seg_re, x):
        seg_file = nipy.load_image(x)
        hgg_seg_list.append(np.array(seg_file.get_data()).tolist())
    if re.search(t1_re, x):
        t1_file = nipy.load_image(x)
        hgg_t1_list.append(np.array(t1_file.get_data()).tolist())
    if re.search(t1ce_re, x):
        t1ce_file = nipy.load_image(x)
        hgg_t1ce_list.append(np.array(t1ce_file.get_data()).tolist())
    if re.search(t2_re, x):
        t2_file = nipy.load_image(x)
        hgg_t2_list.append(np.array(t2_file.get_data()).tolist())
hgg_flair_nparray = np.array(hgg_flair_list)
hgg_seg_nparray = np.array(hgg_seg_list)
hgg_t1_nparray = np.array(hgg_t1_list)
hgg_t1ce_nparray = np.array(hgg_t1ce_list)
hgg_t2_nparray = np.array(hgg_t2_list)

# LGG labeled data
lgg_folder_path = 'D:\数据\MICCAI_BraTS17_Data_Training_for_TGuo_2017modified\Brats17TrainingData\LGG';
lgg_file_path_list = walkdir(lgg_folder_path, topdown=True);
lgg_flair_list = []
lgg_seg_list = []
lgg_t1_list = []
lgg_t1ce_list =[]
lgg_t2_list = []
for x in lgg_file_path_list:
    if re.search(flair_re,x):
        flair_file = nipy.load_image(x)
        lgg_flair_list.append(np.array(flair_file.get_data()).tolist())
    if re.search(seg_re, x):
        seg_file = nipy.load_image(x)
        lgg_seg_list.append(np.array(seg_file.get_data()).tolist())
    if re.search(t1_re, x):
        t1_file = nipy.load_image(x)
        lgg_t1_list.append(np.array(t1_file.get_data()).tolist())
    if re.search(t1ce_re, x):
        t1ce_file = nipy.load_image(x)
        lgg_t1ce_list.append(np.array(t1ce_file.get_data()).tolist())
    if re.search(t2_re, x):
        t2_file = nipy.load_image(x)
        lgg_t2_list.append(np.array(t2_file.get_data()).tolist())
lgg_flair_nparray = np.array(lgg_flair_list)
lgg_seg_nparray =   np.array(lgg_seg_list)
lgg_t1_nparray =    np.array(lgg_t1_list)
lgg_t1ce_nparray =  np.array(lgg_t1ce_list)
lgg_t2_nparray =    np.array(lgg_t2_list)




dset_train = f_train['data']
label_train = f_train['label']
dset_val = f_val['data']
label_val = f_val['label']
dset_test = f_test['data']

# extract features from data
keys = ['T2W', 'ADC', 'DWI1', 'DWI2', 'DWI3', 'ROI']
radiomics_parameter_path = r'RadiomicsParams.yaml'
extractor = featureextractor.RadiomicsFeaturesExtractor(radiomics_parameter_path)

# save to excel data
wb = Workbook()
ws = wb.active
column_index = 1
row_num = 0

# ROI = dset_train[0, :, :, 5]
# ROI = np.uint16(ROI[..., np.newaxis])
# ROI = sitk.GetImageFromArray(ROI)
# image0 = dset_train[0, :, :, 0]
# image0 = image0[..., np.newaxis]
# image00 = sitk.GetImageFromArray(image0)
# demo0 = extractor.execute(image00, ROI)
# image1 = dset_train[0, :, :, 1]
# image1 = image1[..., np.newaxis]
# image11= sitk.GetImageFromArray(image1)
# demo1 = extractor.execute(image11, ROI)
#
# for item0, item1 in zip(demo0.items(), demo1.items()):
#     print(item0, item1)
# def same (a, b):
#     if a==b:
#         return True
# print(same(image00,image11))

def XLRef(row, column, zero_indexed=True):
    if zero_indexed:
        row += 1
        column += 1
    return get_column_letter(column) + str(row)

dset_train = dset_test
for num in range(dset_train.shape[0]):
    for x in range(len(keys)-1):

        ROI = dset_train[num, :, :, 5]
        ROI = np.uint16(ROI[..., np.newaxis])
        ROI = sitk.GetImageFromArray(ROI)
        image = dset_train[num, :, :, x]
        image = image[..., np.newaxis]
        image = sitk.GetImageFromArray(image)
        demo = extractor.execute(image, ROI)
        key = [x for x in demo]
        vals = [x[1] for x in demo.items()]
        # ws[XLRef(row_num * len(vals) + 1, column_index)] = keys[row_num]
        for index in range(len(vals)):
            if None == ws[XLRef(index + len(vals) * x, 0)].value:
                ws[XLRef(index + len(vals) * x, 0)] = key[index]
            else:
                assert (ws[XLRef(index + len(vals) * x, 0)].value == key[index])
            if isinstance(vals[index], numbers.Number):
                ws[XLRef(index+ len(vals) * x, column_index)] = vals[index]
            else:
                ws[XLRef(index+ len(vals) * x, column_index)] = str(vals[index])

    # ws[XLRef(570, 0)] = 'label'
    # ws[XLRef(570, num+1)] = int(label_train[num][0])
    column_index += 1
wb.save(r'new new data test all features .xlsx')
