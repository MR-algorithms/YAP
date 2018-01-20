from radiomics import featureextractor
import numpy as np
## import h5py
import SimpleITK as sitk
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import numbers

f_train = h5py.File(r'Z:\Radiomics_ZhangJing\ProstateX\new_data\training_2D_101x101_box_center.h5')
f_test = h5py.File(r'Z:\Radiomics_ZhangJing\ProstateX\new_data\testing_2D_101x101_box_center_no_label.h5')
f_val = h5py.File(r'Z:\Radiomics_ZhangJing\ProstateX\new_data\validation_2D_101x101_box_center.h5')

dset_train = f_train['data']
label_train = f_train['label']
dset_val = f_val['data']
label_val = f_val['label']
dset_test = f_test['data']

# extract features from data
keys = ['T2W', 'ADC', 'DWI1', 'DWI2', 'DWI3', 'ROI']
radiomics_parameter_path = r'RadiomicsParams.yaml'
extractor = featureextractor.RadiomicsFeaturesExtractor(radiomics_parameter_path)

# save to excel data
wb = Workbook()
ws = wb.active
column_index = 1
row_num = 0

# ROI = dset_train[0, :, :, 5]
# ROI = np.uint16(ROI[..., np.newaxis])
# ROI = sitk.GetImageFromArray(ROI)
# image0 = dset_train[0, :, :, 0]
# image0 = image0[..., np.newaxis]
# image00 = sitk.GetImageFromArray(image0)
# demo0 = extractor.execute(image00, ROI)
# image1 = dset_train[0, :, :, 1]
# image1 = image1[..., np.newaxis]
# image11= sitk.GetImageFromArray(image1)
# demo1 = extractor.execute(image11, ROI)
#
# for item0, item1 in zip(demo0.items(), demo1.items()):
#     print(item0, item1)
# def same (a, b):
#     if a==b:
#         return True
# print(same(image00,image11))

def XLRef(row, column, zero_indexed=True):
    if zero_indexed:
        row += 1
        column += 1
    return get_column_letter(column) + str(row)

dset_train = dset_test
for num in range(dset_train.shape[0]):
    for x in range(len(keys)-1):

        ROI = dset_train[num, :, :, 5]
        ROI = np.uint16(ROI[..., np.newaxis])
        ROI = sitk.GetImageFromArray(ROI)
        image = dset_train[num, :, :, x]
        image = image[..., np.newaxis]
        image = sitk.GetImageFromArray(image)
        demo = extractor.execute(image, ROI)
        key = [x for x in demo]
        vals = [x[1] for x in demo.items()]
        # ws[XLRef(row_num * len(vals) + 1, column_index)] = keys[row_num]
        for index in range(len(vals)):
            if None == ws[XLRef(index + len(vals) * x, 0)].value:
                ws[XLRef(index + len(vals) * x, 0)] = key[index]
            else:
                assert (ws[XLRef(index + len(vals) * x, 0)].value == key[index])
            if isinstance(vals[index], numbers.Number):
                ws[XLRef(index+ len(vals) * x, column_index)] = vals[index]
            else:
                ws[XLRef(index+ len(vals) * x, column_index)] = str(vals[index])

    # ws[XLRef(570, 0)] = 'label'
    # ws[XLRef(570, num+1)] = int(label_train[num][0])
    column_index += 1
wb.save(r'new new data test all features .xlsx')
