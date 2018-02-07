from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import numbers

wb = Workbook()


def XL_Location(row, column):
    return get_column_letter(column) + str(row)


def Save_Column_Title(file_dir, features, row_index, column_start):
    ws = wb.active
    keys = [x[0] for x in features.items()]
    try:
        for index in range(len(keys)):
            loc = XL_Location(index + row_index, column_start)
            if ws[loc].value is None:
                ws[loc] = keys[index]
            else:
                assert (ws[loc].value == keys[index])
        wb.save(file_dir)
    except:
        return False
    return True


# save to excel data
def Save_Column_Exel(file_dir, features, row_index, column_start):
    ws = wb.active
    vals = [x[1] for x in features.items()]
    try:
        for index in range(len(vals)):
            loc = XL_Location(index + row_index, column_start)
            if isinstance(vals[index], numbers.Number):
                ws[XL_Location(index + row_index, column_start)] = vals[index]
            else:
                ws[XL_Location(index + row_index, column_start)] = str(vals[index])
        wb.save(file_dir)
    except:
        return False
    return True


def Save_Row_Exel(file_dir, features, row_index, column_start):
    ws = wb.active
    vals = [x[1] for x in features.items()]
    try:
        for index in range(len(vals)):
            loc = XL_Location(index + row_index, column_start)
            if isinstance(vals[index], numbers.Number):
                ws[XL_Location(row_index, index + column_start)] = vals[index]
            else:
                ws[XL_Location(row_index, index + column_start)] = str(vals[index])
        wb.save(file_dir)
    except:
        return False
    return True